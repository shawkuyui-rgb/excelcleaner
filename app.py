from __future__ import annotations

import time
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from excel_cleaner import (
    CleanReport,
    CleanupRuleError,
    SheetResult,
    _copy_print_area,
    _normalize_print_ranges,
    _resolve_retained_sheets,
)


SECRET_PASSWORD = "Avaval123"
MAX_UPLOAD_SIZE_BYTES = 100 * 1024 * 1024
READABLE_SIZE_LIMIT = "100MB"


class WorkbookReadError(RuntimeError):
    """Raised when an uploaded workbook cannot be read safely."""


@dataclass(frozen=True)
class MemoryCleanResult:
    output_bytes: bytes
    output_filename: str
    report: CleanReport
    elapsed_seconds: float
    source_size_bytes: int
    source_sheet_count: int


def main() -> None:
    st.set_page_config(page_title="Excel 清理工具", layout="centered")
    _inject_styles()

    if not _password_gate():
        return

    _render_app()


def _password_gate() -> bool:
    if st.session_state.get("authenticated"):
        return True

    with st.form("password_form", clear_on_submit=False):
        password = st.text_input("访问密码", type="password", placeholder="请输入密码")
        submitted = st.form_submit_button("确认进入", use_container_width=True)

    if submitted:
        if password == SECRET_PASSWORD:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("密码错误，请重新输入。")
    return False


def _render_app() -> None:
    st.title("Excel 清理工具")
    st.caption("上传文件，在内存中完成清理，然后直接下载结果文件。")

    uploaded_file = st.file_uploader(
        "拖拽或选择 Excel 文件",
        type=["xlsx"],
        accept_multiple_files=False,
        help=f"仅支持 .xlsx，最大 {READABLE_SIZE_LIMIT}。",
    )

    st.subheader("参数配置")
    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.checkbox("仅保留打印区域内容", value=True, disabled=True)
    with col_b:
        st.checkbox("自动清理隐藏的行与列", value=True, disabled=True)
    with col_c:
        st.checkbox("公式转静态数值", value=True, disabled=True)

    validation_error = _validate_upload(uploaded_file)
    if validation_error:
        st.error(validation_error)

    can_process = uploaded_file is not None and validation_error is None

    st.subheader("执行")
    if st.button("开始清理", type="primary", use_container_width=True, disabled=not can_process):
        assert uploaded_file is not None
        uploaded_bytes = uploaded_file.getvalue()
        with st.spinner("正在处理复杂数据，请勿刷新页面..."):
            try:
                result = clean_workbook_bytes(uploaded_bytes, uploaded_file.name)
            except CleanupRuleError as exc:
                st.error(str(exc))
                return
            except WorkbookReadError as exc:
                st.error(str(exc))
                return
            except Exception as exc:
                st.error(f"处理失败：{exc}")
                return

        st.success("清理完成，结果文件已准备好。")
        _render_receipt(result)
        st.download_button(
            "下载清理后的 Excel 文件",
            data=result.output_bytes,
            file_name=result.output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


def clean_workbook_bytes(uploaded_bytes: bytes, filename: str) -> MemoryCleanResult:
    start_time = time.perf_counter()
    source_name = Path(filename).name
    output_filename = _build_output_filename(source_name)

    try:
        formula_book = load_workbook(BytesIO(uploaded_bytes), data_only=False)
        value_book = load_workbook(BytesIO(uploaded_bytes), data_only=True)
    except (BadZipFile, InvalidFileException, KeyError, RuntimeError, OSError) as exc:
        raise WorkbookReadError("无法读取：请先在 Excel 中解除文件密码保护后再上传。") from exc

    cleaned_book = Workbook()
    cleaned_book.remove(cleaned_book.active)

    retained_sheet_names, removed_by_dnp_names, left_marker, right_marker = _resolve_retained_sheets(formula_book.sheetnames)
    report = CleanReport(
        source_file=Path(source_name),
        output_file=Path(output_filename),
        retained_sheet_names=retained_sheet_names,
        removed_by_dnp_names=removed_by_dnp_names,
        dnp_left_marker=left_marker,
        dnp_right_marker=right_marker,
    )

    for sheet_name in retained_sheet_names:
        source_sheet = formula_book[sheet_name]
        value_sheet = value_book[sheet_name]
        print_ranges = _normalize_print_ranges(source_sheet)

        if not print_ranges:
            report.skipped_sheets.append(
                SheetResult(name=sheet_name, status="skipped", message="No print area defined.")
            )
            continue

        new_sheet = cleaned_book.create_sheet(title=sheet_name)
        report.processed_sheets.append(_copy_print_area(source_sheet, value_sheet, new_sheet, print_ranges))

    if not report.processed_sheets:
        cleaned_book.create_sheet(title="Summary")
        summary_sheet = cleaned_book["Summary"]
        summary_sheet["A1"] = "No sheets were cleaned because no print areas were defined."
        summary_sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFF4CC")

    output_buffer = BytesIO()
    cleaned_book.save(output_buffer)
    output_buffer.seek(0)

    return MemoryCleanResult(
        output_bytes=output_buffer.getvalue(),
        output_filename=output_filename,
        report=report,
        elapsed_seconds=time.perf_counter() - start_time,
        source_size_bytes=len(uploaded_bytes),
        source_sheet_count=len(formula_book.sheetnames),
    )


def _validate_upload(uploaded_file: object | None) -> str | None:
    if uploaded_file is None:
        return None

    filename = getattr(uploaded_file, "name", "")
    if Path(filename).suffix.lower() != ".xlsx":
        return "仅支持 .xlsx 格式文件，请勿上传 .xls、.csv 或其他格式。"

    size = getattr(uploaded_file, "size", None)
    if size is not None and size > MAX_UPLOAD_SIZE_BYTES:
        return "文件超过 100MB 限制"

    return None


def _build_output_filename(filename: str) -> str:
    source = Path(filename)
    stem = source.stem or "cleaned_workbook"
    return f"{stem}_cleaned.xlsx"


def _render_receipt(result: MemoryCleanResult) -> None:
    report = result.report
    receipt_rows = [
        ("处理耗时", f"{result.elapsed_seconds:.2f} 秒"),
        ("清理前文件大小", _format_size(result.source_size_bytes)),
        ("清理后文件大小", _format_size(len(result.output_bytes))),
        ("原始 Sheet 数", str(result.source_sheet_count)),
        ("DNP 保留 Sheet 数", str(len(report.retained_sheet_names))),
        ("DNP 删除 Sheet 数", str(len(report.removed_by_dnp_names))),
        ("实际处理 Sheet 数", str(report.processed_count)),
        ("跳过无打印区域 Sheet 数", str(report.skipped_count)),
        ("清理隐藏行数", str(report.hidden_rows_removed_count)),
        ("清理隐藏列数", str(report.hidden_columns_removed_count)),
        ("公式转空值数", str(report.formula_to_empty_count)),
        ("保留图片数", str(report.images_kept_count)),
    ]

    st.subheader("处理回执")
    metric_cols = st.columns(4)
    metric_cols[0].metric("处理耗时", f"{result.elapsed_seconds:.2f}s")
    metric_cols[1].metric("处理 Sheet", report.processed_count)
    metric_cols[2].metric("跳过 Sheet", report.skipped_count)
    metric_cols[3].metric("隐藏行列", report.hidden_rows_removed_count + report.hidden_columns_removed_count)

    st.table([{"项目": label, "结果": value} for label, value in receipt_rows])

    if report.skipped_sheets:
        skipped_names = "、".join(sheet.name for sheet in report.skipped_sheets)
        st.warning(f"以下 Sheet 因未设置打印区域被跳过：{skipped_names}")


def _format_size(size_bytes: int) -> str:
    units = ("B", "KB", "MB", "GB")
    value = float(size_bytes)
    for unit in units:
        if value < 1024 or unit == units[-1]:
            return f"{value:.1f} {unit}" if unit != "B" else f"{int(value)} B"
        value /= 1024
    return f"{size_bytes} B"


def _inject_styles() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background: #eef4f7;
        }
        .block-container {
            max-width: 980px;
            padding-top: 3rem;
            padding-bottom: 4rem;
        }
        .auth-shell {
            max-width: 520px;
            margin: 12vh auto 0;
            padding: 2rem;
            border: 1px solid #dce8ee;
            border-radius: 8px;
            background: #ffffff;
        }
        h1, h2, h3 {
            color: #102631;
            letter-spacing: 0;
        }
        div[data-testid="stFileUploader"] section {
            border-color: #b8cbd4;
            background: #ffffff;
        }
        div[data-testid="stMetric"] {
            background: #ffffff;
            border: 1px solid #dce8ee;
            border-radius: 8px;
            padding: 1rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
