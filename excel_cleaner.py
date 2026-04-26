from __future__ import annotations

import copy
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet


class CleanupRuleError(ValueError):
    """Raised when the workbook does not satisfy cleanup preconditions."""


@dataclass
class SheetResult:
    name: str
    status: str
    message: str = ""
    copied_cells: int = 0
    formula_to_empty: int = 0
    hidden_rows_removed: int = 0
    hidden_columns_removed: int = 0
    images_kept: int = 0


@dataclass
class CleanReport:
    source_file: Path
    output_file: Path
    processed_sheets: List[SheetResult] = field(default_factory=list)
    skipped_sheets: List[SheetResult] = field(default_factory=list)
    retained_sheet_names: List[str] = field(default_factory=list)
    removed_by_dnp_names: List[str] = field(default_factory=list)
    dnp_left_marker: str = ""
    dnp_right_marker: str = ""

    @property
    def processed_count(self) -> int:
        return len(self.processed_sheets)

    @property
    def skipped_count(self) -> int:
        return len(self.skipped_sheets)

    @property
    def formula_to_empty_count(self) -> int:
        return sum(item.formula_to_empty for item in self.processed_sheets)

    @property
    def hidden_rows_removed_count(self) -> int:
        return sum(item.hidden_rows_removed for item in self.processed_sheets)

    @property
    def hidden_columns_removed_count(self) -> int:
        return sum(item.hidden_columns_removed for item in self.processed_sheets)

    @property
    def images_kept_count(self) -> int:
        return sum(item.images_kept for item in self.processed_sheets)


def clean_workbook(source_path: str | Path, output_dir: str | Path | None = None) -> CleanReport:
    source = Path(source_path).expanduser().resolve()
    if source.suffix.lower() != ".xlsx":
        raise ValueError("Only .xlsx files are supported.")
    if not source.exists():
        raise FileNotFoundError(f"File not found: {source}")

    destination_dir = Path(output_dir).expanduser().resolve() if output_dir else source.parent / "cleaned_output"
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = _build_output_path(source, destination_dir)

    formula_book = load_workbook(source, data_only=False)
    value_book = load_workbook(source, data_only=True)
    cleaned_book = Workbook()
    cleaned_book.remove(cleaned_book.active)

    retained_sheet_names, removed_by_dnp_names, left_marker, right_marker = _resolve_retained_sheets(formula_book.sheetnames)
    report = CleanReport(
        source_file=source,
        output_file=destination,
        retained_sheet_names=retained_sheet_names,
        removed_by_dnp_names=removed_by_dnp_names,
        dnp_left_marker=left_marker,
        dnp_right_marker=right_marker,
    )

    for sheet_name in retained_sheet_names:
        source_sheet = formula_book[sheet_name]
        value_sheet = value_book[sheet_name]
        print_ranges = _normalize_print_ranges(source_sheet)

        if not print_ranges:
            report.skipped_sheets.append(
                SheetResult(name=sheet_name, status="skipped", message="No print area defined.")
            )
            continue

        new_sheet = cleaned_book.create_sheet(title=sheet_name)
        result = _copy_print_area(source_sheet, value_sheet, new_sheet, print_ranges)
        report.processed_sheets.append(result)

    if not report.processed_sheets:
        cleaned_book.create_sheet(title="Summary")
        summary_sheet = cleaned_book["Summary"]
        summary_sheet["A1"] = "No sheets were cleaned because no print areas were defined."
        summary_sheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFF4CC")

    cleaned_book.save(destination)
    return report


def _build_output_path(source: Path, output_dir: Path) -> Path:
    candidate = output_dir / f"{source.stem}_cleaned.xlsx"
    counter = 1
    while candidate.exists():
        candidate = output_dir / f"{source.stem}_cleaned_{counter}.xlsx"
        counter += 1
    return candidate


def _resolve_retained_sheets(sheet_names: List[str]) -> Tuple[List[str], List[str], str, str]:
    left_markers = [name for name in sheet_names if "<<DNP" in name]
    right_markers = [name for name in sheet_names if "DNP>>" in name]

    if len(left_markers) != 1:
        raise CleanupRuleError("DNP 规则错误：必须存在且仅存在 1 个名称包含 <<DNP 的 worksheet。")
    if len(right_markers) != 1:
        raise CleanupRuleError("DNP 规则错误：必须存在且仅存在 1 个名称包含 DNP>> 的 worksheet。")

    left_marker = left_markers[0]
    right_marker = right_markers[0]
    left_index = sheet_names.index(left_marker)
    right_index = sheet_names.index(right_marker)

    if left_index >= right_index:
        raise CleanupRuleError("DNP 规则错误：<<DNP 必须位于 DNP>> 左侧。")

    retained = sheet_names[left_index + 1 : right_index]
    removed = [name for idx, name in enumerate(sheet_names) if idx <= left_index or idx >= right_index]
    return retained, removed, left_marker, right_marker


def _normalize_print_ranges(sheet: Worksheet) -> List[Tuple[int, int, int, int]]:
    if not sheet.print_area:
        return []

    raw_ranges = sheet.print_area
    if isinstance(raw_ranges, str):
        raw_ranges = [part.strip() for part in raw_ranges.split(",") if part.strip()]

    normalized: List[Tuple[int, int, int, int]] = []
    for item in raw_ranges:
        cleaned = item.replace(f"'{sheet.title}'!", "").replace(f"{sheet.title}!", "")
        min_col, min_row, max_col, max_row = range_boundaries(cleaned)
        normalized.append((min_row, max_row, min_col, max_col))
    return normalized


def _copy_print_area(
    source_sheet: Worksheet,
    value_sheet: Worksheet,
    new_sheet: Worksheet,
    print_ranges: Iterable[Tuple[int, int, int, int]],
) -> SheetResult:
    print_ranges = list(print_ranges)
    row_set = set()
    col_set = set()
    hidden_row_set = set()
    hidden_col_set = set()

    for min_row, max_row, min_col, max_col in print_ranges:
        for row_idx in range(min_row, max_row + 1):
            if source_sheet.row_dimensions[row_idx].hidden:
                hidden_row_set.add(row_idx)
                continue
            row_set.add(row_idx)
        for col_idx in range(min_col, max_col + 1):
            letter = get_column_letter(col_idx)
            if source_sheet.column_dimensions[letter].hidden:
                hidden_col_set.add(col_idx)
                continue
            col_set.add(col_idx)

    if _should_keep_cover_images(source_sheet.title):
        for image in getattr(source_sheet, "_images", []):
            row_idx, col_idx = _get_image_anchor_position(image)
            if row_idx is None or col_idx is None:
                continue
            if not source_sheet.row_dimensions[row_idx].hidden:
                row_set.add(row_idx)
            if not source_sheet.column_dimensions[get_column_letter(col_idx)].hidden:
                col_set.add(col_idx)

    rows_to_keep = sorted(row_set)
    cols_to_keep = sorted(col_set)

    row_map = {original: new for new, original in enumerate(rows_to_keep, start=1)}
    col_map = {original: new for new, original in enumerate(cols_to_keep, start=1)}

    copied_cells = 0
    formula_to_empty = 0

    for min_row, max_row, min_col, max_col in print_ranges:
        for row_idx in range(min_row, max_row + 1):
            if row_idx not in row_map:
                continue
            for col_idx in range(min_col, max_col + 1):
                if col_idx not in col_map:
                    continue
                source_cell = source_sheet.cell(row=row_idx, column=col_idx)
                if _is_empty_cell(source_cell):
                    continue
                target_cell = new_sheet.cell(row=row_map[row_idx], column=col_map[col_idx])
                copied_cells += 1
                formula_to_empty += _copy_cell(source_cell, value_sheet.cell(row=row_idx, column=col_idx), target_cell)

    _copy_dimensions(source_sheet, new_sheet, row_map, col_map)
    _copy_merges(source_sheet, new_sheet, row_map, col_map)
    _copy_sheet_view(source_sheet, new_sheet, row_map, col_map)
    images_kept = _copy_images(source_sheet, new_sheet, row_map, col_map)

    print_area_rows = [row_map[row_idx] for min_row, max_row, _, _ in print_ranges for row_idx in range(min_row, max_row + 1) if row_idx in row_map]
    print_area_cols = [col_map[col_idx] for _, _, min_col, max_col in print_ranges for col_idx in range(min_col, max_col + 1) if col_idx in col_map]
    if print_area_rows and print_area_cols:
        new_sheet.print_area = (
            f"{get_column_letter(min(print_area_cols))}{min(print_area_rows)}:"
            f"{get_column_letter(max(print_area_cols))}{max(print_area_rows)}"
        )

    return SheetResult(
        name=source_sheet.title,
        status="processed",
        message="Cleaned successfully.",
        copied_cells=copied_cells,
        formula_to_empty=formula_to_empty,
        hidden_rows_removed=len(hidden_row_set),
        hidden_columns_removed=len(hidden_col_set),
        images_kept=images_kept,
    )


def _copy_cell(source_cell: Cell, value_cell: Cell, target_cell: Cell) -> int:
    formula_to_empty = 0

    if source_cell.data_type == "f":
        target_cell.value = value_cell.value
        if value_cell.value is None:
            formula_to_empty = 1
    else:
        target_cell.value = source_cell.value

    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.font:
        target_cell.font = copy.copy(source_cell.font)
    if source_cell.fill:
        target_cell.fill = copy.copy(source_cell.fill)
    if source_cell.border:
        target_cell.border = copy.copy(source_cell.border)
    if source_cell.alignment:
        target_cell.alignment = copy.copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy.copy(source_cell.protection)
    if source_cell.hyperlink:
        target_cell._hyperlink = copy.copy(source_cell.hyperlink)

    return formula_to_empty


def _copy_dimensions(source_sheet: Worksheet, target_sheet: Worksheet, row_map: Dict[int, int], col_map: Dict[int, int]) -> None:
    for old_row, new_row in row_map.items():
        source_dim = source_sheet.row_dimensions[old_row]
        target_dim = target_sheet.row_dimensions[new_row]
        target_dim.height = source_dim.height
        target_dim.hidden = False

    for old_col, new_col in col_map.items():
        source_letter = get_column_letter(old_col)
        target_letter = get_column_letter(new_col)
        source_dim = source_sheet.column_dimensions[source_letter]
        target_dim = target_sheet.column_dimensions[target_letter]
        target_dim.width = source_dim.width
        target_dim.hidden = False


def _copy_merges(source_sheet: Worksheet, target_sheet: Worksheet, row_map: Dict[int, int], col_map: Dict[int, int]) -> None:
    for merged_range in source_sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        if not all(row in row_map for row in range(min_row, max_row + 1)):
            continue
        if not all(col in col_map for col in range(min_col, max_col + 1)):
            continue

        mapped = (
            row_map[min_row],
            row_map[max_row],
            col_map[min_col],
            col_map[max_col],
        )
        target_sheet.merge_cells(
            start_row=mapped[0],
            end_row=mapped[1],
            start_column=mapped[2],
            end_column=mapped[3],
        )


def _copy_sheet_view(source_sheet: Worksheet, target_sheet: Worksheet, row_map: Dict[int, int], col_map: Dict[int, int]) -> None:
    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines

    freeze_ref = source_sheet.freeze_panes
    if not freeze_ref:
        return

    if isinstance(freeze_ref, str):
        probe = source_sheet[freeze_ref]
    else:
        probe = freeze_ref

    row_idx = probe.row
    col_idx = probe.column
    if row_idx in row_map and col_idx in col_map:
        target_sheet.freeze_panes = target_sheet.cell(row=row_map[row_idx], column=col_map[col_idx])


def _copy_images(
    source_sheet: Worksheet,
    target_sheet: Worksheet,
    row_map: Dict[int, int],
    col_map: Dict[int, int],
) -> int:
    if not _should_keep_cover_images(source_sheet.title):
        return 0

    kept = 0
    for image in getattr(source_sheet, "_images", []):
        row_idx, col_idx = _get_image_anchor_position(image)
        if row_idx is None or col_idx is None:
            continue
        if row_idx not in row_map or col_idx not in col_map:
            continue

        payload = _read_image_bytes(image)
        if payload is None:
            continue

        new_image = Image(BytesIO(payload))
        new_image.width = image.width
        new_image.height = image.height
        anchor = target_sheet.cell(row=row_map[row_idx], column=col_map[col_idx]).coordinate
        target_sheet.add_image(new_image, anchor)
        kept += 1

    return kept


def _get_image_anchor_position(image: Image) -> Tuple[int | None, int | None]:
    anchor = image.anchor
    if isinstance(anchor, str):
        return coordinate_to_tuple(anchor)

    marker = getattr(anchor, "_from", None)
    if marker is None:
        return None, None
    return marker.row + 1, marker.col + 1


def _read_image_bytes(image: Image) -> bytes | None:
    ref = image.ref
    if hasattr(ref, "getvalue"):
        return ref.getvalue()
    if hasattr(ref, "read"):
        current_pos = ref.tell() if hasattr(ref, "tell") else None
        if hasattr(ref, "seek"):
            ref.seek(0)
        payload = ref.read()
        if current_pos is not None and hasattr(ref, "seek"):
            ref.seek(current_pos)
        return payload
    if isinstance(ref, (str, Path)):
        return Path(ref).read_bytes()
    return None


def _should_keep_cover_images(sheet_title: str) -> bool:
    return "cover" in sheet_title.lower()


def _is_empty_cell(cell: Cell) -> bool:
    return cell.value is None and not cell.has_style
